from __future__ import annotations

import argparse
import math
import re
from collections import defaultdict
from datetime import date, datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


NAVY = "16324F"
BLUE = "1F4E78"
CYAN = "D9EAF7"
TEAL = "0F766E"
GREEN = "D9EAD3"
PALE_GREEN = "E2F0D9"
AMBER = "FCE5CD"
RED = "F4CCCC"
PALE_RED = "FCE8E6"
GRAY = "E7E6E6"
PALE_GRAY = "F4F6F8"
WHITE = "FFFFFF"
TEXT = "1F2937"
THIN_GRAY = Side(style="thin", color="D1D5DB")
MONEY_FMT = '"S/" #,##0.00;[Red]-"S/" #,##0.00'
NUMBER_FMT = '#,##0;[Red]-#,##0'
DECIMAL_FMT = '#,##0.0;[Red]-#,##0.0'
PERCENT_FMT = '0.0%;[Red]-0.0%'
PP_FMT = '0.0%;[Red]-0.0%'

MONTHS = ("Junio", "Julio")
WEEKDAY_ORDER = {
    "lunes": 0,
    "martes": 1,
    "miercoles": 2,
    "miércoles": 2,
    "jueves": 3,
    "viernes": 4,
    "sabado": 5,
    "sábado": 5,
    "domingo": 6,
}


def num(value) -> float:
    try:
        parsed = float(value)
        return parsed if math.isfinite(parsed) else 0.0
    except (TypeError, ValueError):
        return 0.0


def ratio(numerator: float, denominator: float) -> float:
    return numerator / denominator if denominator else 0.0


def parse_date(value) -> date:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    return datetime.strptime(str(value)[:10], "%Y-%m-%d").date()


def event_month(event_name: str) -> str | None:
    text = str(event_name or "").upper()
    if "JUNIO 2026" in text:
        return "Junio"
    if "JULIO 2026" in text:
        return "Julio"
    return None


def rating(occupancy: float) -> str:
    if occupancy < 0.05:
        return "CRÍTICO"
    if occupancy < 0.15:
        return "DÉBIL"
    if occupancy < 0.40:
        return "SALUDABLE"
    return "FUERTE"


def hour_key(label: str):
    match = re.search(r"(\d{2}):(\d{2})", str(label))
    return (int(match.group(1)), int(match.group(2)), str(label)) if match else (99, 99, str(label))


def read_table(ws):
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(value) if value is not None else "" for value in rows[0]]
    return [dict(zip(headers, values)) for values in rows[1:] if any(value is not None for value in values)]


def aggregate(rows):
    dates = {row["_date"] for row in rows}
    result = {
        "dates": len(dates),
        "slots": len(rows),
        "capacity": sum(num(row.get("capacidad")) for row in rows),
        "visits": sum(num(row.get("visitas_ocupadas")) for row in rows),
        "direct": sum(num(row.get("entradas_directas")) for row in rows),
        "bag": sum(num(row.get("reservas_bolsa")) for row in rows),
        "attended": sum(num(row.get("visitas_asistidas")) for row in rows),
        "courtesy": sum(num(row.get("cortesias_o_valor_cero")) for row in rows),
        "revenue": sum(num(row.get("ingreso_atribuido")) for row in rows),
        "zero_slots": sum(1 for row in rows if num(row.get("visitas_ocupadas")) == 0),
    }
    result["visits_per_date"] = ratio(result["visits"], result["dates"])
    result["visits_per_slot"] = ratio(result["visits"], result["slots"])
    result["occupancy"] = ratio(result["visits"], result["capacity"])
    result["attendance"] = ratio(result["attended"], result["visits"])
    result["zero_rate"] = ratio(result["zero_slots"], result["slots"])
    result["revenue_per_visit"] = ratio(result["revenue"], result["visits"])
    result["weak_slots"] = sum(1 for row in rows if ratio(num(row.get("visitas_ocupadas")), num(row.get("capacidad"))) < 0.15)
    result["critical_slots"] = sum(1 for row in rows if ratio(num(row.get("visitas_ocupadas")), num(row.get("capacidad"))) < 0.05)
    return result


def group_aggregate(rows, key_fn):
    groups = defaultdict(list)
    for row in rows:
        groups[key_fn(row)].append(row)
    return {key: aggregate(values) for key, values in groups.items()}


def safe_variation(current, previous):
    return (current - previous) / previous if previous else None


def table_name(value: str) -> str:
    clean = re.sub(r"[^A-Za-z0-9_]", "", value)
    return f"T_{clean[:220]}"


def section(ws, row: int, col_start: int, col_end: int, text: str):
    ws.merge_cells(start_row=row, start_column=col_start, end_row=row, end_column=col_end)
    cell = ws.cell(row, col_start, text)
    cell.fill = PatternFill("solid", fgColor=BLUE)
    cell.font = Font(color=WHITE, bold=True, size=11)
    cell.alignment = Alignment(vertical="center")
    ws.row_dimensions[row].height = 22


def title(ws, text: str, subtitle: str, last_col: int):
    ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=last_col)
    cell = ws.cell(1, 1, text)
    cell.fill = PatternFill("solid", fgColor=NAVY)
    cell.font = Font(color=WHITE, bold=True, size=20)
    cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 12
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=last_col)
    cell = ws.cell(3, 1, subtitle)
    cell.font = Font(color="4B5563", italic=True, size=10)
    cell.alignment = Alignment(wrap_text=True, vertical="center")
    ws.row_dimensions[3].height = 28


def write_table(ws, start_row, start_col, headers, rows, name, formats=None, widths=None):
    formats = formats or {}
    widths = widths or {}
    for index, header in enumerate(headers, start_col):
        cell = ws.cell(start_row, index, header)
        cell.fill = PatternFill("solid", fgColor=NAVY)
        cell.font = Font(color=WHITE, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(bottom=THIN_GRAY)
    for row_offset, values in enumerate(rows, start_row + 1):
        for col_offset, value in enumerate(values, start_col):
            cell = ws.cell(row_offset, col_offset, value)
            cell.border = Border(bottom=THIN_GRAY)
            cell.alignment = Alignment(vertical="center", wrap_text=col_offset == start_col)
            key = headers[col_offset - start_col]
            if key in formats:
                cell.number_format = formats[key]
        if (row_offset - start_row) % 2 == 0:
            for col_offset in range(start_col, start_col + len(headers)):
                ws.cell(row_offset, col_offset).fill = PatternFill("solid", fgColor=PALE_GRAY)
    end_row = start_row + max(1, len(rows))
    end_col = start_col + len(headers) - 1
    if rows:
        ref = f"{ws.cell(start_row, start_col).coordinate}:{ws.cell(end_row, end_col).coordinate}"
        tab = Table(displayName=table_name(name), ref=ref)
        tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True, showColumnStripes=False)
        ws.add_table(tab)
    for idx, header in enumerate(headers, start_col):
        ws.column_dimensions[get_column_letter(idx)].width = widths.get(header, max(12, min(30, len(header) + 3)))
    return end_row, end_col


def add_weak_format(ws, cell_range):
    ws.conditional_formatting.add(
        cell_range,
        CellIsRule(operator="lessThan", formula=["0.05"], fill=PatternFill("solid", fgColor=RED)),
    )
    ws.conditional_formatting.add(
        cell_range,
        CellIsRule(operator="between", formula=["0.05", "0.149999"], fill=PatternFill("solid", fgColor=AMBER)),
    )
    ws.conditional_formatting.add(
        cell_range,
        CellIsRule(operator="greaterThanOrEqual", formula=["0.15"], fill=PatternFill("solid", fgColor=PALE_GREEN)),
    )


def setup_sheet(ws, landscape=True):
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A5"
    ws.auto_filter = None
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_setup.orientation = "landscape" if landscape else "portrait"
    ws.page_margins.left = 0.25
    ws.page_margins.right = 0.25
    ws.page_margins.top = 0.5
    ws.page_margins.bottom = 0.5
    ws.oddFooter.center.text = "Página &P de &N"
    ws.oddFooter.right.text = "FDNDA · Piscina Libre"


def main():
    parser = argparse.ArgumentParser(description="Genera el comparativo de Piscina Libre junio vs julio 2026")
    parser.add_argument("source", type=Path, help="Reporte comercial XLSX fuente")
    parser.add_argument("output", type=Path, help="Archivo XLSX de salida")
    args = parser.parse_args()

    source = args.source.resolve()
    output = args.output.resolve()
    source_wb = load_workbook(source, read_only=True, data_only=True)
    detail = read_table(source_wb["Detalle"])
    products = read_table(source_wb["Ventas producto"])
    source_summary = read_table(source_wb["Resumen"])
    summary_map = {str(row.get("indicador")): row.get("valor") for row in source_summary}
    cutoff_text = str(summary_map.get("Corte Lima", "2026-07-31 00:00:00"))
    cutoff = parse_date(cutoff_text)

    selected_detail = []
    operational = {month: [] for month in MONTHS}
    for row in detail:
        month = event_month(row.get("evento"))
        if month is None:
            continue
        prepared = dict(row)
        prepared["_month"] = month
        prepared["_date"] = parse_date(row.get("fecha"))
        prepared["_offered_completed"] = (
            str(row.get("habilitado", "")).strip().upper() == "SI"
            and num(row.get("capacidad")) > 0
            and prepared["_date"] < cutoff
        )
        selected_detail.append(prepared)
        if prepared["_offered_completed"]:
            operational[month].append(prepared)

    selected_products = []
    sales = {month: {"units": 0.0, "gross": 0.0, "discount": 0.0, "effective": 0.0} for month in MONTHS}
    for row in products:
        month = event_month(row.get("evento"))
        if month is None:
            continue
        prepared = dict(row)
        prepared["_month"] = month
        selected_products.append(prepared)
        sales[month]["units"] += num(row.get("unidades"))
        sales[month]["gross"] += num(row.get("ingreso_bruto_lista"))
        sales[month]["discount"] += num(row.get("descuentos"))
        sales[month]["effective"] += num(row.get("ingreso_efectivo"))
    for month in MONTHS:
        sales[month]["effective_per_unit"] = ratio(sales[month]["effective"], sales[month]["units"])
        sales[month]["discount_rate"] = ratio(sales[month]["discount"], sales[month]["gross"])

    totals = {month: aggregate(operational[month]) for month in MONTHS}
    hours = {month: group_aggregate(operational[month], lambda row: row.get("horario")) for month in MONTHS}
    weekdays = {month: group_aggregate(operational[month], lambda row: str(row.get("dia")).lower()) for month in MONTHS}
    dates = {month: group_aggregate(operational[month], lambda row: row["_date"]) for month in MONTHS}
    day_hours = {
        month: group_aggregate(operational[month], lambda row: (str(row.get("dia")).lower(), row.get("horario")))
        for month in MONTHS
    }

    wb = Workbook()
    wb.remove(wb.active)
    wb.properties.title = "Comparativo Piscina Libre Junio vs Julio 2026"
    wb.properties.subject = "Ingresos, variación comercial, días y horarios débiles"
    wb.properties.creator = "FDNDA"
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True

    # Resumen ejecutivo
    ws = wb.create_sheet("Resumen ejecutivo")
    setup_sheet(ws)
    title(
        ws,
        "PISCINA LIBRE · JUNIO VS JULIO 2026",
        f"Corte de datos: {cutoff_text} (America/Lima). Operación comparable: 1–30 de junio vs 1–30 de julio; el 31/07 se excluye por estar parcial.",
        12,
    )
    section(ws, 5, 1, 5, "Resultado comercial por evento mensual")
    headers = ["Indicador", "Junio", "Julio", "Diferencia", "Variación %"]
    dashboard_rows = [
        ["Ingreso efectivo", sales["Junio"]["effective"], sales["Julio"]["effective"], sales["Julio"]["effective"] - sales["Junio"]["effective"], safe_variation(sales["Julio"]["effective"], sales["Junio"]["effective"])],
        ["Ingreso bruto a lista", sales["Junio"]["gross"], sales["Julio"]["gross"], sales["Julio"]["gross"] - sales["Junio"]["gross"], safe_variation(sales["Julio"]["gross"], sales["Junio"]["gross"])],
        ["Descuentos", sales["Junio"]["discount"], sales["Julio"]["discount"], sales["Julio"]["discount"] - sales["Junio"]["discount"], safe_variation(sales["Julio"]["discount"], sales["Junio"]["discount"])],
        ["Unidades vendidas", sales["Junio"]["units"], sales["Julio"]["units"], sales["Julio"]["units"] - sales["Junio"]["units"], safe_variation(sales["Julio"]["units"], sales["Junio"]["units"])],
    ]
    write_table(
        ws, 6, 1, headers, dashboard_rows, "ResumenComercial",
        formats={"Junio": MONEY_FMT, "Julio": MONEY_FMT, "Diferencia": MONEY_FMT, "Variación %": PERCENT_FMT},
        widths={"Indicador": 27, "Junio": 16, "Julio": 16, "Diferencia": 17, "Variación %": 14},
    )
    for cell in (ws["B10"], ws["C10"], ws["D10"]):
        cell.number_format = NUMBER_FMT

    section(ws, 12, 1, 5, "Operación comparable por fecha de servicio")
    operation_rows = [
        ["Visitas ocupadas", totals["Junio"]["visits"], totals["Julio"]["visits"], totals["Julio"]["visits"] - totals["Junio"]["visits"], safe_variation(totals["Julio"]["visits"], totals["Junio"]["visits"])],
        ["Ingreso atribuido a visitas", totals["Junio"]["revenue"], totals["Julio"]["revenue"], totals["Julio"]["revenue"] - totals["Junio"]["revenue"], safe_variation(totals["Julio"]["revenue"], totals["Junio"]["revenue"])],
        ["Ocupación", totals["Junio"]["occupancy"], totals["Julio"]["occupancy"], totals["Julio"]["occupancy"] - totals["Junio"]["occupancy"], safe_variation(totals["Julio"]["occupancy"], totals["Junio"]["occupancy"])],
        ["Asistencia registrada", totals["Junio"]["attendance"], totals["Julio"]["attendance"], totals["Julio"]["attendance"] - totals["Junio"]["attendance"], safe_variation(totals["Julio"]["attendance"], totals["Junio"]["attendance"])],
    ]
    write_table(
        ws, 13, 1, headers, operation_rows, "ResumenOperacion",
        formats={"Junio": NUMBER_FMT, "Julio": NUMBER_FMT, "Diferencia": NUMBER_FMT, "Variación %": PERCENT_FMT},
        widths={"Indicador": 27, "Junio": 16, "Julio": 16, "Diferencia": 17, "Variación %": 14},
    )
    for coord in ("B15", "C15", "D15"):
        ws[coord].number_format = MONEY_FMT
    for row_number in (16, 17):
        for column in range(2, 6):
            ws.cell(row_number, column).number_format = PERCENT_FMT

    weakest = {}
    for month in MONTHS:
        eligible_hours = [(key, value) for key, value in hours[month].items() if value["dates"] >= 2]
        weakest_hour = min(eligible_hours, key=lambda item: (item[1]["occupancy"], hour_key(item[0])))
        weakest_day = min(weekdays[month].items(), key=lambda item: item[1]["occupancy"])
        weakest_date = min(dates[month].items(), key=lambda item: (item[1]["occupancy"], item[0]))
        weakest[month] = (weakest_hour, weakest_day, weakest_date)

    section(ws, 5, 7, 12, "Lectura ejecutiva")
    gain = sales["Julio"]["effective"] - sales["Junio"]["effective"]
    gain_pct = safe_variation(sales["Julio"]["effective"], sales["Junio"]["effective"])
    ws.merge_cells("G6:L7")
    ws["G6"] = f"Julio registra S/ {gain:,.2f} más de ingreso efectivo que junio ({gain_pct:.1%}). Esta es la ganancia comparativa de ventas, no utilidad contable."
    ws["G6"].fill = PatternFill("solid", fgColor=GREEN if gain >= 0 else RED)
    ws["G6"].font = Font(color=TEXT, bold=True, size=12)
    ws["G6"].alignment = Alignment(wrap_text=True, vertical="center")
    ws.merge_cells("G9:L10")
    visit_change = safe_variation(totals["Julio"]["visits"], totals["Junio"]["visits"])
    occ_pp = totals["Julio"]["occupancy"] - totals["Junio"]["occupancy"]
    ws["G9"] = f"En el tramo operativo comparable, las visitas cambian {visit_change:+.1%} y la ocupación {occ_pp:+.1%} frente a junio."
    ws["G9"].fill = PatternFill("solid", fgColor=CYAN)
    ws["G9"].font = Font(color=TEXT, bold=True)
    ws["G9"].alignment = Alignment(wrap_text=True, vertical="center")

    for month, start_row in (("Junio", 12), ("Julio", 15)):
        (hour_label, hour_data), (day_label, day_data), (date_value, date_data) = weakest[month]
        ws.merge_cells(start_row=start_row, start_column=7, end_row=start_row + 1, end_column=12)
        cell = ws.cell(start_row, 7)
        cell.value = (
            f"{month}: horario más débil {hour_label} ({hour_data['occupancy']:.1%}); "
            f"día de semana más débil {day_label} ({day_data['occupancy']:.1%}); "
            f"fecha más débil {date_value.strftime('%d/%m/%Y')} ({date_data['occupancy']:.1%})."
        )
        cell.fill = PatternFill("solid", fgColor=AMBER)
        cell.font = Font(color=TEXT, bold=True)
        cell.alignment = Alignment(wrap_text=True, vertical="center")

    ws.merge_cells("G18:L19")
    ws["G18"] = "Nota: sin costos de personal, energía, mantenimiento y comisiones no se calcula utilidad neta. El archivo usa ingreso efectivo y su variación como indicador comercial."
    ws["G18"].fill = PatternFill("solid", fgColor=PALE_RED)
    ws["G18"].alignment = Alignment(wrap_text=True, vertical="center")
    ws["G18"].font = Font(color="7F1D1D", italic=True)

    chart = BarChart()
    chart.type = "col"
    chart.style = 10
    chart.title = "Ingreso efectivo por evento mensual"
    chart.y_axis.title = "Soles"
    chart.height = 7.5
    chart.width = 12
    chart.add_data(Reference(ws, min_col=2, max_col=3, min_row=7, max_row=7), titles_from_data=False, from_rows=True)
    chart.set_categories(Reference(ws, min_col=2, max_col=3, min_row=6, max_row=6))
    ws.add_chart(chart, "G21")
    ws.freeze_panes = "A6"
    ws.print_area = "A1:L36"

    # Comparativo mensual detallado
    ws = wb.create_sheet("Comparativo mensual")
    setup_sheet(ws)
    title(ws, "COMPARATIVO MENSUAL", "Ventas del evento mensual y operación por fecha de servicio. La diferencia siempre es Julio − Junio.", 6)
    metrics = [
        ("VENTAS", "Unidades vendidas", "unidades", sales["Junio"]["units"], sales["Julio"]["units"], NUMBER_FMT),
        ("VENTAS", "Ingreso bruto a tarifa de lista", "S/", sales["Junio"]["gross"], sales["Julio"]["gross"], MONEY_FMT),
        ("VENTAS", "Descuentos aplicados", "S/", sales["Junio"]["discount"], sales["Julio"]["discount"], MONEY_FMT),
        ("VENTAS", "Tasa de descuento", "%", sales["Junio"]["discount_rate"], sales["Julio"]["discount_rate"], PERCENT_FMT),
        ("VENTAS", "Ingreso efectivo", "S/", sales["Junio"]["effective"], sales["Julio"]["effective"], MONEY_FMT),
        ("VENTAS", "Precio efectivo por unidad", "S/", sales["Junio"]["effective_per_unit"], sales["Julio"]["effective_per_unit"], MONEY_FMT),
        ("OPERACIÓN", "Fechas cerradas comparadas", "fechas", totals["Junio"]["dates"], totals["Julio"]["dates"], NUMBER_FMT),
        ("OPERACIÓN", "Turnos ofrecidos", "turnos", totals["Junio"]["slots"], totals["Julio"]["slots"], NUMBER_FMT),
        ("OPERACIÓN", "Capacidad ofrecida", "cupos", totals["Junio"]["capacity"], totals["Julio"]["capacity"], NUMBER_FMT),
        ("OPERACIÓN", "Visitas ocupadas", "visitas", totals["Junio"]["visits"], totals["Julio"]["visits"], NUMBER_FMT),
        ("OPERACIÓN", "Entradas directas", "visitas", totals["Junio"]["direct"], totals["Julio"]["direct"], NUMBER_FMT),
        ("OPERACIÓN", "Reservas de bolsa", "visitas", totals["Junio"]["bag"], totals["Julio"]["bag"], NUMBER_FMT),
        ("OPERACIÓN", "Visitas por fecha", "promedio", totals["Junio"]["visits_per_date"], totals["Julio"]["visits_per_date"], DECIMAL_FMT),
        ("OPERACIÓN", "Ocupación", "%", totals["Junio"]["occupancy"], totals["Julio"]["occupancy"], PERCENT_FMT),
        ("OPERACIÓN", "Asistencia registrada", "%", totals["Junio"]["attendance"], totals["Julio"]["attendance"], PERCENT_FMT),
        ("OPERACIÓN", "Ingreso atribuido a visitas", "S/", totals["Junio"]["revenue"], totals["Julio"]["revenue"], MONEY_FMT),
        ("OPERACIÓN", "Ingreso atribuido por visita", "S/", totals["Junio"]["revenue_per_visit"], totals["Julio"]["revenue_per_visit"], MONEY_FMT),
        ("DEBILIDAD", "Turnos en cero", "%", totals["Junio"]["zero_rate"], totals["Julio"]["zero_rate"], PERCENT_FMT),
        ("DEBILIDAD", "Turnos débiles (<15%)", "turnos", totals["Junio"]["weak_slots"], totals["Julio"]["weak_slots"], NUMBER_FMT),
        ("DEBILIDAD", "Turnos críticos (<5%)", "turnos", totals["Junio"]["critical_slots"], totals["Julio"]["critical_slots"], NUMBER_FMT),
    ]
    comp_rows = []
    formats_by_row = {}
    for idx, (block, metric, unit, june, july, fmt) in enumerate(metrics, 6):
        comp_rows.append([block, metric, unit, june, july, july - june, safe_variation(july, june)])
        formats_by_row[idx] = fmt
    headers = ["Bloque", "Indicador", "Unidad", "Junio", "Julio", "Diferencia", "Variación %"]
    end_row, _ = write_table(ws, 5, 1, headers, comp_rows, "ComparativoMensual", widths={"Bloque": 14, "Indicador": 34, "Unidad": 12, "Junio": 16, "Julio": 16, "Diferencia": 16, "Variación %": 14})
    for row_no, fmt in formats_by_row.items():
        for col_no in (4, 5, 6):
            ws.cell(row_no, col_no).number_format = fmt
        ws.cell(row_no, 7).number_format = PERCENT_FMT
    ws.conditional_formatting.add(f"F6:F{end_row}", ColorScaleRule(start_type="min", start_color=RED, mid_type="percentile", mid_value=50, mid_color=WHITE, end_type="max", end_color=GREEN))
    ws.freeze_panes = "D6"

    # Horarios
    ws = wb.create_sheet("Horarios comparados")
    setup_sheet(ws)
    title(ws, "HORARIOS · JUNIO VS JULIO", "Rojo: crítico (<5%). Ámbar: débil (5%–14.9%). Verde: saludable o fuerte (≥15%).", 21)
    all_hours = sorted(set(hours["Junio"]) | set(hours["Julio"]), key=hour_key)
    hour_rows = []
    for label in all_hours:
        jn = hours["Junio"].get(label, aggregate([]))
        jl = hours["Julio"].get(label, aggregate([]))
        hour_rows.append([
            label,
            jn["dates"], jn["capacity"], jn["visits"], jn["visits_per_date"], jn["occupancy"], jn["zero_rate"], jn["revenue"], rating(jn["occupancy"]),
            jl["dates"], jl["capacity"], jl["visits"], jl["visits_per_date"], jl["occupancy"], jl["zero_rate"], jl["revenue"], rating(jl["occupancy"]),
            jl["visits_per_date"] - jn["visits_per_date"], jl["occupancy"] - jn["occupancy"], jl["revenue"] - jn["revenue"],
            "MEJORA" if jl["occupancy"] > jn["occupancy"] else "CAÍDA" if jl["occupancy"] < jn["occupancy"] else "IGUAL",
        ])
    headers = ["Horario", "Jun días", "Jun capacidad", "Jun visitas", "Jun prom./día", "Jun ocupación", "Jun ceros", "Jun ingreso", "Jun nivel", "Jul días", "Jul capacidad", "Jul visitas", "Jul prom./día", "Jul ocupación", "Jul ceros", "Jul ingreso", "Jul nivel", "Δ prom./día", "Δ ocupación", "Δ ingreso", "Evolución"]
    end_row, _ = write_table(ws, 5, 1, headers, hour_rows, "HorariosComparados", widths={"Horario": 18, "Jun nivel": 13, "Jul nivel": 13, "Evolución": 12})
    for row_no in range(6, end_row + 1):
        for col_no in (6, 7, 14, 15, 19):
            ws.cell(row_no, col_no).number_format = PERCENT_FMT
        for col_no in (8, 16, 20):
            ws.cell(row_no, col_no).number_format = MONEY_FMT
        for col_no in (5, 13, 18):
            ws.cell(row_no, col_no).number_format = DECIMAL_FMT
    add_weak_format(ws, f"F6:F{end_row}")
    add_weak_format(ws, f"N6:N{end_row}")
    ws.freeze_panes = "B6"

    # Días de semana
    ws = wb.create_sheet("Días comparados")
    setup_sheet(ws)
    title(ws, "DÍAS DE SEMANA · JUNIO VS JULIO", "Comparación ajustada por capacidad de todos los turnos ofrecidos en fechas cerradas.", 19)
    all_days = sorted(set(weekdays["Junio"]) | set(weekdays["Julio"]), key=lambda value: WEEKDAY_ORDER.get(value, 99))
    day_rows = []
    for label in all_days:
        jn = weekdays["Junio"].get(label, aggregate([]))
        jl = weekdays["Julio"].get(label, aggregate([]))
        day_rows.append([
            label.capitalize(),
            jn["dates"], jn["slots"], jn["capacity"], jn["visits"], jn["visits_per_date"], jn["occupancy"], jn["revenue"], rating(jn["occupancy"]),
            jl["dates"], jl["slots"], jl["capacity"], jl["visits"], jl["visits_per_date"], jl["occupancy"], jl["revenue"], rating(jl["occupancy"]),
            jl["occupancy"] - jn["occupancy"], jl["revenue"] - jn["revenue"],
        ])
    headers = ["Día", "Jun fechas", "Jun turnos", "Jun capacidad", "Jun visitas", "Jun visitas/fecha", "Jun ocupación", "Jun ingreso", "Jun nivel", "Jul fechas", "Jul turnos", "Jul capacidad", "Jul visitas", "Jul visitas/fecha", "Jul ocupación", "Jul ingreso", "Jul nivel", "Δ ocupación", "Δ ingreso"]
    end_row, _ = write_table(ws, 5, 1, headers, day_rows, "DiasComparados", widths={"Día": 14, "Jun nivel": 13, "Jul nivel": 13})
    for row_no in range(6, end_row + 1):
        for col_no in (7, 15, 18):
            ws.cell(row_no, col_no).number_format = PERCENT_FMT
        for col_no in (8, 16, 19):
            ws.cell(row_no, col_no).number_format = MONEY_FMT
        for col_no in (6, 14):
            ws.cell(row_no, col_no).number_format = DECIMAL_FMT
    add_weak_format(ws, f"G6:G{end_row}")
    add_weak_format(ws, f"O6:O{end_row}")
    ws.freeze_panes = "B6"

    # Fechas débiles
    ws = wb.create_sheet("Fechas débiles")
    setup_sheet(ws)
    title(ws, "FECHAS MÁS DÉBILES", "Ranking por menor ocupación; solo fechas cerradas y turnos habilitados. Se muestran todas las fechas con ocupación menor a 15%.", 10)
    weak_date_rows = []
    for month in MONTHS:
        ranked = sorted(dates[month].items(), key=lambda item: (item[1]["occupancy"], item[1]["visits"], item[0]))
        rank_no = 0
        for date_value, values in ranked:
            if values["occupancy"] >= 0.15:
                continue
            rank_no += 1
            day_label = next((str(row.get("dia")).capitalize() for row in operational[month] if row["_date"] == date_value), "")
            weak_date_rows.append([month, rank_no, date_value, day_label, values["slots"], values["capacity"], values["visits"], values["occupancy"], values["revenue"], rating(values["occupancy"])])
    headers = ["Mes", "Ranking", "Fecha", "Día", "Turnos", "Capacidad", "Visitas", "Ocupación", "Ingreso atribuido", "Nivel"]
    end_row, _ = write_table(ws, 5, 1, headers, weak_date_rows, "FechasDebiles", formats={"Fecha": "dd/mm/yyyy", "Ocupación": PERCENT_FMT, "Ingreso atribuido": MONEY_FMT}, widths={"Mes": 12, "Ranking": 10, "Fecha": 13, "Día": 14, "Ingreso atribuido": 18, "Nivel": 13})
    add_weak_format(ws, f"H6:H{end_row}")
    ws.freeze_panes = "A6"

    # Día + horario débil
    ws = wb.create_sheet("Día-hora débiles")
    setup_sheet(ws)
    title(ws, "COMBINACIONES DÍA + HORARIO DÉBILES", "Prioridades con al menos 2 ocurrencias y ocupación inferior a 15% en el mes.", 12)
    weak_combo_rows = []
    for month in MONTHS:
        ranked = sorted(day_hours[month].items(), key=lambda item: (item[1]["occupancy"], WEEKDAY_ORDER.get(item[0][0], 99), hour_key(item[0][1])))
        for (day_label, hour_label), values in ranked:
            if values["slots"] < 2 or values["occupancy"] >= 0.15:
                continue
            priority = "ALTA" if values["occupancy"] < 0.05 else "MEDIA"
            weak_combo_rows.append([month, day_label.capitalize(), hour_label, values["slots"], values["capacity"], values["visits"], values["visits_per_slot"], values["occupancy"], values["zero_rate"], values["revenue"], rating(values["occupancy"]), priority])
    headers = ["Mes", "Día", "Horario", "Ocurrencias", "Capacidad", "Visitas", "Prom./turno", "Ocupación", "Turnos en cero", "Ingreso atribuido", "Nivel", "Prioridad"]
    end_row, _ = write_table(ws, 5, 1, headers, weak_combo_rows, "DiaHoraDebiles", formats={"Prom./turno": DECIMAL_FMT, "Ocupación": PERCENT_FMT, "Turnos en cero": PERCENT_FMT, "Ingreso atribuido": MONEY_FMT}, widths={"Mes": 12, "Día": 14, "Horario": 18, "Ingreso atribuido": 18, "Nivel": 13, "Prioridad": 12})
    add_weak_format(ws, f"H6:H{end_row}")
    ws.freeze_panes = "A6"

    # Ventas por producto
    ws = wb.create_sheet("Ventas por producto")
    setup_sheet(ws)
    title(ws, "VENTAS POR PRODUCTO", "Productos pagados asociados a los eventos mensuales de Piscina Libre.", 12)
    product_headers = ["Mes", "Evento", "Producto", "Tipo", "Unidades", "Pedidos*", "Compradores*", "Tarifa lista", "Ingreso bruto", "Descuentos", "Ingreso efectivo", "Precio medio"]
    product_rows = []
    for row in selected_products:
        product_rows.append([
            row["_month"], row.get("evento"), row.get("producto"), row.get("tipo"), num(row.get("unidades")), num(row.get("pedidos")), num(row.get("compradores")), num(row.get("tarifa_lista")), num(row.get("ingreso_bruto_lista")), num(row.get("descuentos")), num(row.get("ingreso_efectivo")), num(row.get("precio_medio_efectivo")),
        ])
    end_row, _ = write_table(ws, 5, 1, product_headers, product_rows, "VentasProductoJJ", formats={"Tarifa lista": MONEY_FMT, "Ingreso bruto": MONEY_FMT, "Descuentos": MONEY_FMT, "Ingreso efectivo": MONEY_FMT, "Precio medio": MONEY_FMT}, widths={"Mes": 12, "Evento": 53, "Producto": 31, "Tipo": 22})
    ws.merge_cells(start_row=end_row + 2, start_column=1, end_row=end_row + 3, end_column=12)
    ws.cell(end_row + 2, 1, "* Pedidos y compradores se muestran por producto; no deben sumarse entre filas porque una misma compra puede incluir más de un producto.")
    ws.cell(end_row + 2, 1).alignment = Alignment(wrap_text=True, vertical="center")
    ws.cell(end_row + 2, 1).fill = PatternFill("solid", fgColor=AMBER)
    ws.freeze_panes = "A6"

    # Detalle base
    ws = wb.create_sheet("Detalle base")
    setup_sheet(ws)
    title(ws, "DETALLE BASE", "Trazabilidad por fecha y horario. 'Incluido comparación' identifica filas usadas para medir días y horarios débiles.", 21)
    detail_headers = ["Mes", "Evento", "Fecha", "Día", "Horario", "Capacidad", "Habilitado", "Incluido comparación", "Vendido inventario", "Entradas directas", "Reservas bolsa", "Visitas ocupadas", "Visitas asistidas", "Cortesías/valor cero", "Ocupación", "Ingreso atribuido", "Tarifa lista", "Anticipación mediana", "Diferencia inventario", "Nivel", "Observación"]
    detail_rows = []
    for row in sorted(selected_detail, key=lambda value: (MONTHS.index(value["_month"]), value["_date"], hour_key(value.get("horario")))):
        occupancy = ratio(num(row.get("visitas_ocupadas")), num(row.get("capacidad")))
        included = row["_offered_completed"]
        observation = "" if included else ("Fecha parcial o futura" if row["_date"] >= cutoff else "Horario deshabilitado o sin capacidad")
        detail_rows.append([
            row["_month"], row.get("evento"), row["_date"], str(row.get("dia")).capitalize(), row.get("horario"), num(row.get("capacidad")), row.get("habilitado"), "SÍ" if included else "NO", num(row.get("vendidos_inventario")), num(row.get("entradas_directas")), num(row.get("reservas_bolsa")), num(row.get("visitas_ocupadas")), num(row.get("visitas_asistidas")), num(row.get("cortesias_o_valor_cero")), occupancy, num(row.get("ingreso_atribuido")), num(row.get("tarifa_lista")), num(row.get("anticipacion_mediana_dias")), num(row.get("diferencia_inventario_vs_demanda")), rating(occupancy), observation,
        ])
    end_row, _ = write_table(ws, 5, 1, detail_headers, detail_rows, "DetalleBaseJJ", formats={"Fecha": "dd/mm/yyyy", "Ocupación": PERCENT_FMT, "Ingreso atribuido": MONEY_FMT, "Tarifa lista": MONEY_FMT}, widths={"Mes": 12, "Evento": 53, "Fecha": 13, "Día": 14, "Horario": 18, "Incluido comparación": 19, "Observación": 30})
    add_weak_format(ws, f"O6:O{end_row}")
    ws.freeze_panes = "F6"

    # Metodología
    ws = wb.create_sheet("Metodología")
    setup_sheet(ws, landscape=False)
    title(ws, "METODOLOGÍA Y LÍMITES", "Definiciones para interpretar correctamente el comparativo.", 3)
    methodology_rows = [
        ["Fuente", "Base de datos de producción; reporte comercial fuente", source.name],
        ["Corte", "Fecha y hora de extracción en America/Lima", cutoff_text],
        ["Meses", "Eventos mensuales incluidos", "Piscina Libre junio 2026 y julio 2026"],
        ["Ingreso bruto", "Suma del valor de productos antes de descuentos", "No equivale a caja neta"],
        ["Ingreso efectivo", "Monto pagado después de descuentos, asignado proporcionalmente a productos", "Principal KPI comercial"],
        ["Ganancia comparativa", "Ingreso efectivo de julio menos ingreso efectivo de junio", "Es variación comercial; no utilidad contable"],
        ["Utilidad neta", "No calculada porque la fuente no contiene todos los costos", "Requiere personal, energía, mantenimiento, comisiones y otros costos"],
        ["Demanda operativa", "Entradas directas ACTIVE/PAID más reservas de bolsa RESERVED/USED", "Medida por fecha de servicio"],
        ["Ingreso atribuido", "Valor efectivo asignado a cada visita; bolsas se distribuyen por créditos", "Puede diferir del ingreso vendido si quedan créditos sin reservar"],
        ["Periodo comparable", "01/06–30/06 frente a 01/07–30/07", "El 31/07 está excluido por ser parcial al corte"],
        ["Horario ofrecido", "Inventario habilitado con capacidad mayor a cero", "Cierres operativos quedan fuera del ranking"],
        ["Crítico", "Ocupación menor a 5%", "Prioridad alta"],
        ["Débil", "Ocupación desde 5% hasta menos de 15%", "Prioridad media"],
        ["Saludable", "Ocupación desde 15% hasta menos de 40%", "Seguimiento"],
        ["Fuerte", "Ocupación de 40% o más", "Proteger precio y redirigir excedentes"],
        ["Asistencia", "Visitas con uso registrado en el sistema", "Puede subestimarse si hubo acceso sin escaneo"],
    ]
    write_table(ws, 5, 1, ["Tema", "Definición", "Interpretación"], methodology_rows, "MetodologiaJJ", widths={"Tema": 24, "Definición": 62, "Interpretación": 62})
    for row_no in range(6, 6 + len(methodology_rows)):
        ws.row_dimensions[row_no].height = 34
    ws.freeze_panes = "A6"

    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is not None and cell.row > 3 and cell.alignment == Alignment():
                    cell.alignment = Alignment(vertical="center")

    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)
    print(f"Archivo generado: {output}")
    print(f"Corte: {cutoff_text}")
    print(f"Ingreso efectivo junio: {sales['Junio']['effective']:.2f}")
    print(f"Ingreso efectivo julio: {sales['Julio']['effective']:.2f}")
    print(f"Diferencia: {gain:.2f} ({gain_pct:.2%})")
    print(f"Visitas junio/julio comparables: {totals['Junio']['visits']:.0f}/{totals['Julio']['visits']:.0f}")


if __name__ == "__main__":
    main()
